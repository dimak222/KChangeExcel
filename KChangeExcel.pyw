#-------------------------------------------------------------------------------
# Author:      dimak222
#
# Created:     15.03.2023
# Copyright:   (c) dimak222 2023
# Licence:     No
#-------------------------------------------------------------------------------

title = "KChangeExcel"
ver = "v1.1.0.0"
url = "https://github.com/dimak222/KChangeExcel" # ссылка на файл

#------------------------------Настройки!---------------------------------------

check_update = True # проверять обновление программы ("True" - да; "False" или "" - нет)
beta = False # скачивать бета версии программы ("True" - да; "False" или "" - нет)

Marking_and_Name_by_source = True # обозначение и наименование всех дет. и СБ по источнику ("True" - да; "False" или "" - нет)

force_processing = True # обрабатывать все дет. (помогает, если дет. с одинаковым обозначением и\или наименованем вставлена в СБ разными файлами, но увеличивает время обработки) ("True" - да; "False" или "" - нет)

property_library = "Свойства СПМ.lpt" # название файла со св-ми (должен лежать в C:\Program Files\ASCON\KOMPAS-3D vХХ\Sys)

#------------------------------Импорт модулей-----------------------------------

import os # работа с файовой системой
import psutil # модуль вывода запущеных процессов

import sys # модуль информации об операционной системе

import pythoncom # модуль для запуска без IDE
from win32com.client import Dispatch, gencache # библиотека API Windows
from sys import exit # для выхода из приложения без ошибки

from threading import Thread # библиотека потоков

import tkinter.messagebox as mb # окно с сообщением
import tkinter as tk # модуль окон

from openpyxl import load_workbook # модуль управлением Excel

import tkinter.ttk as ttk # модуль окон
import time # модуль времени

from pythoncom import VT_EMPTY # для записи пустого св-ва
from win32com.client import VARIANT # для записи пустого св-ва

import re # модуль регулярных выражений

#-------------------------------------------------------------------------------

def DoubleExe(): # проверка на уже запущеное приложение

    global program_directory # значение делаем глобальным

    filename = title + ".exe" # имя запускаемого файла

    list = [] # список найденых путей

    for process in psutil.process_iter(): # перебор всех процессов

        try: # попытаться узнать имя процесса
            proc_name = process.name() # имя процесса

        except psutil.NoSuchProcess: # в случае ошибки
            pass # пропускаем

        else: # если есть имя
            if proc_name == filename: # сравниваем имя
                list.append(process.cwd())
                if len(list) > 2: # если найдено больше двух названий программы (два процесса)
                    Message("Приложение уже запущено!") # сообщение, поверх всех окон и с автоматическим закрытием
                    exit() # выходим из программы

    if list == []: # если путь не указан
        program_directory = "" # ничего не указывааем (будет рядом с программой)
    else: # если путь найден
        program_directory = os.path.dirname(psutil.Process().exe()) # директория файла

def KompasAPI(): # подключение API КОМПАСа

    try: # попытаться подключиться к КОМПАСу

        global KompasConst # значение делаем глобальным
        global KompasAPI7 # значение делаем глобальным
        global iApplication # значение делаем глобальным
        global iKompasObject # значение делаем глобальным
        global iDocuments # значение делаем глобальным
        global iPropertyMng # значение делаем глобальным

        KompasConst3D = gencache.EnsureModule("{2CAF168C-7961-4B90-9DA2-701419BEEFE3}", 0, 1, 0).constants # константа 3D документов
        KompasConst2D = gencache.EnsureModule("{75C9F5D0-B5B8-4526-8681-9903C567D2ED}", 0, 1, 0).constants # константа 2D документов
        KompasConst = gencache.EnsureModule("{75C9F5D0-B5B8-4526-8681-9903C567D2ED}", 0, 1, 0).constants # константа для скрытия вопросов перестроения

        KompasAPI5 = gencache.EnsureModule("{0422828C-F174-495E-AC5D-D31014DBBE87}", 0, 1, 0) # API5 КОМПАСа
        iKompasObject = Dispatch("Kompas.Application.5", None, KompasAPI5.KompasObject.CLSID) # интерфейс API КОМПАС

        KompasAPI7 = gencache.EnsureModule("{69AC2981-37C0-4379-84FD-5DD2F3C0A520}", 0, 1, 0) # API7 КОМПАСа
        iApplication = Dispatch("Kompas.Application.7") # интерфейс приложения КОМПАС-3D.

        iDocuments = iApplication.Documents # интерфейс для открытия документов

        iPropertyMng = KompasAPI7.IPropertyMng(iApplication) # интерфейс менеджера свойств

        if iApplication.Visible == False: # если компас невидимый
            iApplication.Visible = True # сделать КОМПАС-3D видемым

    except: # если не получилось подключиться к КОМПАСу

        Message("КОМПАС-3D не найден!\nУстановите или переустановите КОМПАС-3D!") # сообщение, поверх всех окон с автоматическим закрытием
        exit() # выходим из программы

def Kompas_message(text): # сообщение в окне КОМПАСа если он открыт

    if iApplication.Visible == True: # если компас видимый
        iApplication.MessageBoxEx(text, 'Message:', 64) # сообщение в КОМПАСе

def Message(text = "Ошибка!", counter = 4): # сообщение, поверх всех окон с автоматическим закрытием (текст, время закрытия)

    def Message_Thread(text, counter): # сообщение, поверх всех окон с автоматическим закрытием (текст, время закрытия)

        if counter == 0: # время до закрытия окна (если 0)
            counter = 1 # закрытие через 1 сек
        window_msg = tk.Tk() # создание окна
        try: # попытаться использовать значёк
            window_msg.iconbitmap(default = Resource_path("cat.ico")) # значёк программы
        except: # если ошибка
            pass # пропустить
        window_msg.attributes("-topmost", True) # окно поверх всех окон
        window_msg.withdraw() # скрываем окно "невидимое"
        time = counter * 1000 # время в милисекундах
        window_msg.after(time, window_msg.destroy) # закрытие окна через n милисекунд
        if mb.showinfo(title, text, parent = window_msg) == "": # информационное окно закрытое по времени
            pass # пропустить
        else: # если не закрыто по времени
            window_msg.destroy() # окно закрыто по кнопке
        window_msg.mainloop() # отображение окна

    msg_th = Thread(target = Message_Thread, args = (text, counter)) # запуск окна в отдельном потоке
    msg_th.start() # запуск потока

    msg_th.join() # ждать завершения процесса, иначе может закрыться следующие окно

def Resource_path(relative_path): # для сохранения картинки внутри exe файла

    try: # попытаться определить путь к папке
        base_path = sys._MEIPASS # путь к временной папки PyInstaller

    except Exception: # если ошибка
        base_path = os.path.abspath(".") # абсолютный путь

    return os.path.join(base_path, relative_path) # объеденяем и возващаем полный путь

def AskYesNoCancel(text): # вопросительное сообщение, поверх всех окон

    ask = tk.Tk() # создание окна
    try: # попытаться использовать значёк
        ask.iconbitmap(default = Resource_path("cat.ico")) # значёк программы
    except: # если ошибка
        pass # пропустить
    ask.attributes("-topmost", True) # окно поверх всех окон
    ask.withdraw() # скрываем окно "невидимое"
    answer = mb.askyesnocancel(title, text) # задаём вопрос
    ask.destroy() # закрываем окно
    ask.mainloop() # отображение окна

    if answer == True or False: # если ответ "Да" или "Нет"
        return answer # возвращаем результат вопроса

    elif answer == None: # если нажали отмена или крестик
        if iApplication.Visible == False: # если компас невидимый
            iApplication.Quit() # закрываем его
        exit() # выходим из программы

def SystemPath(): # определяем папку системных файлов

    global iSystemPath # значение делаем глобальным

    iSystemPath = iKompasObject.ksSystemPath(0) # определяем папку системных файлов

def Settings_from_Excel(): # настройки считанные с Excel

    global check_update # значение делаем глобальным
    global beta # значение делаем глобальным
    global Marking_and_Name_by_source # значение делаем глобальным
    global force_processing # значение делаем глобальным
    global property_library # значение делаем глобальным

    name_txt_file = os.path.join(program_directory, title + ".xlsx") # путь к файлу Excel

    if os.path.exists(name_txt_file): # если есть txt файл использовать его

        try: # попытаться подключиться к файлу Excel

            workbook = load_workbook(name_txt_file) # подключимся к Excel

        except: # если лист Excel не найден

            Message("Файл Excel повреждён!") # сообщение, поверх всех окон с автоматическим закрытием
            exit() # выходим из программы

        try: # попытаться подключиться к файлу Excel

            ws_name = "Настройки" # название листа
            ws = workbook[ws_name] # выберем лист

            for parameter in ws.iter_rows(min_col = 1, min_row = 1, max_col = 2, max_row = ws.max_row, values_only = True): # обойдём все строки и колонки Excel

                if parameter[1].find("True") != -1: # если есть параметр со словом True, обрабатываем его
                    parameter[1] == True # присвоем значение "True"

                elif parameter[1].find("False") != -1 or parameter[1].strip() == "": # если есть параметр со словом False или "", обрабатываем его
                    parameter[1] == False # присвоем значение "False"

                dict_Settings[parameter[0]] = parameter[1] # словарь параметров

            try: # попытаться присвоить параметры с Excel

                check_update = dict_Settings["check_update"] # проверять обновление программы
                beta = dict_Settings["beta"] # скачивать бета версии программы
                Marking_and_Name_by_source = dict_Settings["Marking_and_Name_by_source"] # обозначение и наименование всех дет. и СБ по источнику
                force_processing = dict_Settings["force_processing"] # обработка всех дет.
                property_library = dict_Settings["property_library"] # название файла со св-ми

            except KeyError as e: # если нет параметра выдать сообщение
                Message(f"В Excel отсутствует параметр: {e}, этот и последующие будут использоваться по умолчанию.") # сообщение, поверх всех окон с автоматическим закрытием (текст, время закрытия)

        except: # если лист Excel не найден
            Message("В Excel не найден лист \"Настройки\"\nИспользуються настройки по умолчанию") # сообщение, поверх всех окон с автоматическим закрытием

    else:
        Message("Файл Excel не найден: " + name_txt_file + "\n Положите его рядом с программой!", 8) # сообщение, поверх всех окон с автоматическим закрытием
        exit() # выходим из программы

def CheckUpdate(): # проверить обновление приложение

    if check_update: # если проверка обновлений включена

        try: # попытаться импортировать модуль обновления

            from Updater import Updater # импортируем модуль обновления

            Updater.Update(title, ver, beta, url, Resource_path("cat.ico")) # проверяем обновление (имя программы, версия программы, скачивать бета версию, ссылка на программу)

        except: # не удалось
            pass # пропустить

def Connecting_to_Excel(): # подключение EXcel и изменение св-в

    global name_txt_file # значение делаем глобальным
    global workbook  # значение делаем глобальным
    global ws # значение делаем глобальным

    name_txt_file = os.path.join(program_directory, title + ".xlsx") # путь к файлу Excel

    if os.path.exists(name_txt_file): # если есть txt файл использовать его

        try: # попытаться подключиться к файлу Excel

            workbook = load_workbook(name_txt_file) # подключимся к Excel

        except: # если лист Excel не найден

            Message("Файл Excel повреждён!") # сообщение, поверх всех окон с автоматическим закрытием
            exit() # выходим из программы

        try: # попытаться подключиться к файлу Excel

            ws_name = "v1.0" # название листа
            ws = workbook[ws_name] # выберем лист

        except: # если лист Excel не найден

            Message("Лист в файле Excel отсутствует или имеет несоответствующию версию!", 8) # сообщение, поверх всех окон с автоматическим закрытием
            exit() # выходим из программы

        row_number = 1 # отчёт с 1-ой строки

        for row in ws.iter_rows(min_col = 2, min_row = 2, max_col = ws.max_column, max_row = ws.max_row, values_only = True): # обойдём все строки и колонки Excel

            row_number += 1 # отчёт количества обработаных строк

            if row[0] != None or row[1] != None: # если есть запись старого обозначения или наименования

                list_Excel.append([row, row_number]) # записать колонку в список

                Reset_column(row_number) # сброс изменений из последней колонки если нет значений

            else: # нет записей

                Reset_column(row_number) # сброс изменений из последней колонки если нет значений

        Save_book() # попытаться сохранить книгу

    else:
        Message("Файл Excel не найден: " + name_txt_file + "\n Положите его рядом с программой!", 8) # сообщение, поверх всех окон с автоматическим закрытием
        exit() # выходим из программы

def Reset_column(row_number): # сброс изменений из последней колонки если нет значений

    cell = ws.cell(row = row_number, column = ws.max_column) # указание колонки

    ivalue = cell.value # значение ячейки

    if ivalue != None: # если ячейка не пустая

        cell.value = "" # запись значения в колонку
        cell.hyperlink = None # гиперссылка на файл

def Save_book(): # попытаться сохранить книгу

    try: # попытаться сохранить записанные значения

        workbook.save(name_txt_file) # сохраняем Excel

    except: # если файл Excel открыт

        if AskYesNoCancel("Для работы колонки \"Изменено\" необходимо закрыть Excel.\nДа - повторить попытку;\nНет - продолжить обработку с открытым Excel;"): # вопросительное сообщение, поверх всех окон
            Save_book() # попытаться сохранить книгу

def Сhecking_open_files(): # проверка открытых файлов

    iKompasDocument = iApplication.ActiveDocument # получить текущий активный документ

    if iKompasDocument == None or iKompasDocument.DocumentType not in (4, 5, 7): # если не открыт док. или не 2D док., выдать сообщение (1-чертёж; 2- фрагмент; 3-СП; 4-модель; 5-СБ; 6-текст. док.; 7-тех. СБ;)

        Message("Откройте дет. или СБ!") # сообщение, поверх всех окон с автоматическим закрытием
        exit() # выходим из программы

def Main_assembly(): # обработка главной СБ или дет.

    global Stop # для чтения в потоке

    iKompasDocument = iApplication.ActiveDocument # получить текущий активный документ

    iKompasDocument3D = KompasAPI7.IKompasDocument3D(iKompasDocument) # базовый класс документов-моделей КОМПАС
    iPart7 = iKompasDocument3D.TopPart # интерфейс компонента 3D документа (сам документ)

    iEmbodimentsManager = KompasAPI7.IEmbodimentsManager(iPart7) # интерфейс менеджера исполнений

    iCurrentEmbodimentIndex = iEmbodimentsManager.CurrentEmbodimentIndex # индекс исполнени

    parameters = iPart7.Marking, iPart7.Name, iCurrentEmbodimentIndex, iPart7.FileName # список параметров файла (Обозначение, Наименование, индекс исполнения, путь к файлу)

    list_files.append(parameters) # добавляем в список

    if iPart7.Detail: # если это дет.

        Сhecking_match(iKompasDocument) # проверка на совпадение обозначения или наименования (не открывать/закрывать файл)

    else: # если это СБ

        if Marking_and_Name_by_source: # обозначение и наименование всех дет. и СБ по источнику

            msg_th = Message_processing("Проверка св-в \"По источнику\"!") # выдача сообщений о количестве файлов (сообщение)

            time.sleep(0.1) # что бы успеть запустить поток при малом количестве файлов

        Collect_sources(iPart7) # рекурсивный сбор дет. и СБ (интерфейс компонента 3D документа)

        if Marking_and_Name_by_source: # обозначение и наименование всех дет. и СБ по источнику

            Stop = True # триггер остановки обработки и сообщения

            msg_th.join() # ждать завершения процесса, иначе может закрыться следующие окно

        Сhecking_match(iKompasDocument) # проверка на совпадение обозначения или наименования (не открывать/закрывать файл)

        iKompasDocument3D.RebuildDocument() # перестроить СБ
        iKompasDocument3D.Save() # сохранить изменения

def Message_processing(msg = "Идёт обработка файлов!"): # выдача сообщений о количестве файлов (сообщение)

    global Stop # глобальный параметр остановки сообщения
    global file_number # для чтения в потоке
    global current_file_name # для чтения в потоке

    def Message_count_Thread(msg): # сообщений о количестве файлов в потоке

        global Stop # глобальный параметр остановки обработки

        class ToolTip(object): # отображает подсказку к виджету

            def __init__(self, widget, text):
                self.widget = widget
                self.text = text
                self.acid = None
                self.tipwindow = None
                self.widget.bind('<Enter>', self.enter)
                self.widget.bind('<Leave>', self.leave)
                self.widget.bind('<ButtonRelease>', self.leave)
                self.widget.bind('<Key>', self.leave)

            def enter(self, event):
                self.schedule()

            def leave(self, event):
                self.unschedule()
                self.hidetip()

            def schedule(self):
                self.unschedule()
                self.acid = self.widget.after(300, self.showtip) # через сколько милисунд отображать подсказку

            def unschedule(self):
                idac = self.acid
                if idac:
                    self.widget.after_cancel(idac)
                self.acid = None

            def showtip(self):
                tw = self.tipwindow = tk.Toplevel(self.widget)
                tw.wm_overrideredirect(1)
                tw.wm_attributes('-topmost', 1) # поверх всех окон
                tw.wm_geometry('+%d+%d' % (self.widget.winfo_rootx(), self.widget.winfo_rooty() + self.widget.winfo_height() + 2))
                tk.Label(tw, text = current_file_name, justify = 'left', bg = '#f2f2f2', relief = 'solid', bd = 1, font = "Verdana 10").pack() # положение, цвет и шрифт текста

            def hidetip(self):
                tw = self.tipwindow
                if tw:
                    tw.destroy()
                self.tipwindow = None

        def Update_text(): # обновление отчёта цифр

            def Updating_text(): # обновление текста

                if Stop: # если триггер остановки обработки и сообщения включён
                    print("Остановил поток!")
                    window.quit() # закрываем окно

                else: # триггер выключен
                    text.config(text = str(file_number)) # обновляем текст
                    text.after(300, Updating_text) # через милисекунды запускаем функцию заново

            Updating_text() # обновление текста

        def Button_exit(): # кнопка "Отмена"
            window.quit() # закрываем окно

        window = tk.Tk() # создание окна
        try: # попытаться использовать значёк
            window.iconbitmap(default = Resource_path("cat.ico")) # значёк программы
        except: # если ошибка
            pass # пропустить
        window.title(title) # заголовок окна
        window.attributes("-topmost",True) # окно поверх всех окон
        x = (window.winfo_screenwidth() - window.winfo_reqwidth()) / 2 # положение по центру монитора
        y = (window.winfo_screenheight() - window.winfo_reqheight()) / 2 # положение по центру монитора
        window.wm_geometry("+%d+%d" % (x - 15, y + 40)) # положение по центру монитора
        window.resizable(width = False, height = False) # блокировка изменение размера окна

        f_top = tk.Frame(window) # блок окна (вверх)
        f_top.pack(expand = True, fill = "both") # размещение блока (с возможностью расширяться и заполненем окна во всех направлениях)

        text = tk.Label(f_top, justify=tk.LEFT, font = "Verdana 10", text = msg) # текст в окне
        text.pack(padx = 5, pady = 2) # размещение блока

        text = tk.Label(f_top, fg="green", justify=tk.LEFT, padx = 3, pady = 3, font = "Verdana 10") # текст
        ToolTip(text, current_file_name) # имя текущего файла в виде всплывающего окна
        Update_text() # обновление отчёта цифр
        text.pack() # размещение блока

        button = tk.Button(f_top, font = "Verdana 11", command = Button_exit, text = "Отмена") # действие кнопки
        button.pack(side = "bottom", pady = 3) # размещение блока

        window.mainloop() # отображение окна

        Stop = True # триггер остановки обработки и сообщения

    Stop = False # триггер остановки сообщения (для работы сообщений при повторном вызове)
    file_number = 0 # отчёт от 0-го файла
    current_file_name = "" # что бы избежать ошибки окна в потоке

    msg_th = Thread(target = Message_count_Thread, args = (msg, )) # запуск сообщений о количестве файлов в отдельном потоке
    msg_th.start() # запуск потока

    return msg_th # возращаем запущеный поток для определения его завершения

def Collect_sources(iPart7): # рекурсивный сбор дет. и СБ

    global Stop # для чтения в потоке
    global file_number # для чтения в потоке
    global current_file_name # для чтения в потоке

    iPartsEx = iPart7.PartsEx(0) # список компонентов, включённыхв расчёт (0 - все компоненты (включая копии из операций копирования); 1 - первые экземпляры вставок компонентов (ksPart7CollectionTypeEnum))

    if iPartsEx: # если есть компоненты в сборке

        for iPart7 in iPartsEx: # проверяем каждый элемент из вставленных в СБ

            if iPart7.Standard: # если это стандартная дет.
                print("Стандартная дет.!", iPart7.FileName)
                continue

            elif iPart7.IsLocal: # если это локальная СБ пропускаем её
                print("Локальный компонент!", iPart7.FileName)
                continue

            elif iPart7.IsBillet: # если это заготовка пропускаем её
                print("Вставлена заготовка!", iPart7.FileName)
                continue

            elif iPart7.IsLayoutGeometry: # если это компоновочная геометрия
                print("Компоновочная геометрия!", iPart7.FileName)
                continue

            if Marking_and_Name_by_source: # обозначение и наименование всех дет. и СБ по источнику

                if Stop == False: # если не нажата кнопка "Отмена" или крестик

                    file_number += 1 # отчёт количества обработаных файлов
                    current_file_name = os.path.basename(iPart7.FileName) # имя документа с расширением для вывода названия файла в окно сообщений

                    BySource(iPart7) # простановка св-в по источнику

            iEmbodimentsManager = KompasAPI7.IEmbodimentsManager(iPart7) # интерфейс менеджера исполнений

            iCurrentEmbodimentIndex = iEmbodimentsManager.CurrentEmbodimentIndex # индекс исполнени

            parameters = iPart7.Marking, iPart7.Name, iCurrentEmbodimentIndex, iPart7.FileName # список параметров файла (Обозначение, Наименование, индекс исполнения, путь к файлу)

            if parameters not in list_files: # если нет в списке пути к файлу, добавить (исключает добавление одной и той же детали (одинаковый путь к детали))
                list_files.append(parameters) # добавляем в список

            if not iPart7.Detail: # если это СБ
                Collect_sources(iPart7) # рекурсивный сбор уникальных документов

    else: # если ошибка (нет дет.)
        print("Пустая сборка!")
        pass

def Сhecking_match(iKompasDocument): # проверка на совпадение обозначения или наименования (не открывать/закрывать файл)

    global file_number # для чтения в потоке
    global current_file_name # для чтения в потоке
    global Stop # для чтения в потоке

    file_number = 0 # отчёт от 0-го файла
    Stop = False # триггер остановки сообщения (для работы если не вызвано окно)

    all_failes_number = len(list_Excel) # количество всех строк в списке

    if all_failes_number > 3: # если файлов больше n

        msg_th = Message_count(all_failes_number, "Идёт обработка файлов!") # выдача сообщений о количестве файлов (количество всех файлов, сообщение) + file_number (номер обрабатываемого файла) + current_file_name (текущее название файла)

        time.sleep(0.1) # что бы успеть запустить поток при малом количестве файлов

    iApplication.HideMessage = KompasConst.ksHideMessageNo # скрыть сообщение перестроения и не перестраивать

    for row in list_Excel: # проверяем каждую строчку считаную с Excel

        file_number += 1 # отчёт количества обработаных файлов

        if Stop == False: # если не нажата кнопка "Отмена" или крестик

            iMarking_old = str(row[0][0]).strip() # старое обозначение с удалением пробелов по бокам
            iMarking_old = MarkingEmbodimentDocCode(iMarking_old) # кортеж старого обозначения (Обозначение, исполнение, пробел перед кодом док., код док.)
            iMarking_old = iMarking_old[0] + iMarking_old[1] # обозначение с исп.

            iName_old = str(row[0][1]).strip() # старое наименование с удалением пробелов по бокам

            not_main_assembly = False # показатель гл. СБ
            force_processing_trigger = False # триггер найденого док.

            for file in list_files: # проверяем каждого файла на совпадение

                iMarking = file[0].strip() # обозначение документа с удалением пробелов по бокам
                iMarking = MarkingEmbodimentDocCode(iMarking) # кортеж считанного обозначения (Обозначение, исполнение, пробел перед кодом док., код док.)
                iMarking = iMarking[0] + iMarking[1] # обозначение с исп.

                iName = file[1].strip() # наименование документа с удалением пробелов по бокам

                if iMarking_old != "None" and iMarking_old != "" and iName_old != "None" and iName_old != "": # если старый обозначение и наименование оба записанны

                    if iMarking_old == iMarking and iName_old == iName: # если старое и новое обозначение и наименование совпадают

                        current_file_name = os.path.basename(file[3]) # имя документа с расширением для вывода названия файла в окно сообщений

                        if not_main_assembly: # если не главная СБ или одна дет.
                            iKompasDocument = iDocuments.Open(file[3], False, False) # Открытие файлов (False - в невидимом режиме, False - с возможностью редактирования)
                        else: # если это гл. СБ
                            iKompasDocument = iApplication.ActiveDocument # получить текущий активный документ

                        if Сhange_properties(row[0], file[2], iKompasDocument): # изменение св-в документов (список значений из Excel, индекс исполнения файла, интерфейс документа)

                            iKompasDocument.Save() # iKompasDocument.Close(1) без iKompasDocument.Save() почему-то не работает

                            Change_or_not(row[1], "Да", file[3]) # запись изменения в последнюю колонку (номер строчки, значение, ссылка)

                        else: # если нет изменений
                            Change_or_not(row[1], "Нет", file[3]) # запись изменения в последнюю колонку (номер строчки, значение, ссылка)

                        if not_main_assembly: # если не главная СБ или одна дет.
                            iKompasDocument.Close(1) # 0 - закрыть документ без сохранения; 1 - закрыть документ, сохранив  изменения; 2 - выдать запрос на сохранение документа, если он изменен.

                        not_main_assembly = True # показатель гл. СБ
                        force_processing_trigger = True # триггер найденого док.

                        if not force_processing: # если включена принудительная обработка
                            break # прерываем цикл

                elif iMarking_old == iMarking or iMarking_old == "None" and iName_old == iName or iMarking_old == "" and iName_old == iName: # если обозначение или наименование пустое и наименование совпало

                    current_file_name = os.path.basename(file[3]) # имя документа с расширением для вывода названия файла в окно сообщений

                    if not_main_assembly: # если не главная СБ или одна дет.
                        iKompasDocument = iDocuments.Open(file[3], False, False) # Открытие файлов (False - в невидимом режиме, False - с возможностью редактирования)
                    else: # если это гл. СБ
                        iKompasDocument = iApplication.ActiveDocument # получить текущий активный документ

                    if Сhange_properties(row[0], file[2], iKompasDocument): # изменение св-в документов (список значений из Excel, индекс исполнения файла, интерфейс документа)

                        iKompasDocument.Save() # iKompasDocument.Close(1) без iKompasDocument.Save() почему-то не работает

                        Change_or_not(row[1], "Да", file[3]) # запись изменения в последнюю колонку (номер строчки, значение, ссылка)

                    else: # если нет изменений
                        Change_or_not(row[1], "Нет", file[3]) # запись изменения в последнюю колонку (номер строчки, значение, ссылка)

                    if not_main_assembly: # если не главная СБ или одна дет.
                        iKompasDocument.Close(1) # 0 - закрыть документ без сохранения; 1 - закрыть документ, сохранив  изменения; 2 - выдать запрос на сохранение документа, если он изменен.

                    not_main_assembly = True # показатель гл. СБ
                    force_processing_trigger = True # триггер найденого док.

                    if not force_processing: # если включена принудительная обработка
                        break # прерываем цикл

                not_main_assembly = True # показатель гл. СБ

            else: # если пройден весь цикл
                if not force_processing_trigger: # триггер найденого док.
                    Change_or_not(row[1], "Не найдено") # запись изменения в последнюю колонку

        else: # если нажали кнопку "Отмена" или крестик
            print("Остановили окном!")
            break # прерываем цикл

    iApplication.HideMessage = KompasConst.ksShowMessage # показывать сообщение перестроения

    if all_failes_number > 3: # если файлов больше n

        Stop = True # триггер остановки обработки и сообщения

        msg_th.join() # ждать завершения процесса, иначе может закрыться следующие окно

    try: # попытаться сохранить записанные значения
        workbook.save(name_txt_file) # сохраняем Excel

    except: # если файл Excel открыт
        pass

def BySource(iPart7): # простановка св-в по источнику

    iPropertyKeeper = KompasAPI7.IPropertyKeeper(iPart7) # интерфейс получения/редактирования значения свойств

    iKompasDocument = iApplication.ActiveDocument # получить текущий активный документ

    for iProperty in ["Обозначение", "Наименование"]: # проверка св-в по источнику

        iProperty = iPropertyMng.GetProperty(iKompasDocument, iProperty) # интерфейс свойства

        iPropertyValue = iPropertyKeeper.GetPropertyValue(iProperty, False, True) # получить значение св-ва (интерфейс св-ва, значение св-ва, единици измерения (СИ))

        if not iPropertyValue[2]: # если не по источнику

            print("Св-во \"По источнику\":", iPropertyValue[1])

            iSetPropertyValue = iPropertyKeeper.SetPropertyValue(iProperty, VARIANT(VT_EMPTY, None), True) # установить значение св-ва (интерфейс св-ва, значение св-ва, единици измерения (СИ))
            iProperty.Update() # применим сво-ва

def Message_count(all_failes_number, msg = "Идёт обработка файлов!"): # выдача сообщений о количестве файлов (количество всех файлов, сообщение) + file_number (номер обрабатываемого файла) + current_file_name (текущее название файла)

    global Stop # глобальный параметр остановки сообщения
    global file_number # для чтения в потоке
    global current_file_name # для чтения в потоке

    def Message_count_Thread(all_failes_number, msg): # сообщений о количестве файлов в потоке

        global Stop # глобальный параметр остановки обработки

        class ToolTip(object): # отображает подсказку к виджету

            def __init__(self, widget, text):
                self.widget = widget
                self.text = text
                self.acid = None
                self.tipwindow = None
                self.widget.bind('<Enter>', self.enter)
                self.widget.bind('<Leave>', self.leave)
                self.widget.bind('<ButtonRelease>', self.leave)
                self.widget.bind('<Key>', self.leave)

            def enter(self, event):
                self.schedule()

            def leave(self, event):
                self.unschedule()
                self.hidetip()

            def schedule(self):
                self.unschedule()
                self.acid = self.widget.after(300, self.showtip) # через сколько милисунд отображать подсказку

            def unschedule(self):
                idac = self.acid
                if idac:
                    self.widget.after_cancel(idac)
                self.acid = None

            def showtip(self):
                tw = self.tipwindow = tk.Toplevel(self.widget)
                tw.wm_overrideredirect(1)
                tw.wm_attributes('-topmost', 1) # поверх всех окон
                tw.wm_geometry('+%d+%d' % (self.widget.winfo_rootx(), self.widget.winfo_rooty() + self.widget.winfo_height() + 2))
                tk.Label(tw, text = current_file_name, justify = 'left', bg = '#f2f2f2', relief = 'solid', bd = 1, font = "Verdana 10").pack() # положение, цвет и шрифт текста

            def hidetip(self):
                tw = self.tipwindow
                if tw:
                    tw.destroy()
                self.tipwindow = None

        def Update_text(): # обновление отчёта цифр

            def Updating_text(): # обновление текста

                if Stop: # если триггер остановки обработки и сообщения включён
                    print("Остановил поток!")
                    window.quit() # закрываем окно

                else: # триггер выключен
                    text.config(text = str(file_number) + "/" + str(all_failes_number)) # обновляем текст
                    text.after(300, Updating_text) # через милисекунды запускаем функцию заново

            Updating_text() # обновление текста

        def Update_progress(): # обновление прогресса

            def Updating_progress(): # обновление прогресса

                percent_file_number = percent_all_failes_number * file_number # процент выполнения

                if Stop: # если триггер остановки обработки и сообщения включён
                    print("Остановил поток прогресса!")
                    window.quit() # закрываем окно

                else: # триггер выключен
                    progress['value'] = percent_file_number # процент выполнения
                    window.update() # (update_idletasks не сбрасывет дпока не дошёл до конца)
                    progress.after(300, Updating_progress) # через милисекунды запускаем функцию заново

            Updating_progress() # обновление прогресса

        def Button_exit(): # кнопка "Отмена"
            window.quit() # закрываем окно

        window = tk.Tk() # создание окна
        try: # попытаться использовать значёк
            window.iconbitmap(default = Resource_path("cat.ico")) # значёк программы
        except: # если ошибка
            pass # пропустить
        window.title(title) # заголовок окна
        window.attributes("-topmost",True) # окно поверх всех окон
        x = (window.winfo_screenwidth() - window.winfo_reqwidth()) / 2 # положение по центру монитора
        y = (window.winfo_screenheight() - window.winfo_reqheight()) / 2 # положение по центру монитора
        window.wm_geometry("+%d+%d" % (x - 25, y + 38)) # положение по центру монитора
        window.resizable(width = False, height = False) # блокировка изменение размера окна

        f_top = tk.Frame(window) # блок окна (вверх)
        f_top.pack(expand = True, fill = "both") # размещение блока (с возможностью расширяться и заполненем окна во всех направлениях)

        text = tk.Label(f_top, justify=tk.LEFT, font = "Verdana 10", text = msg) # текст в окне
        text.pack(padx = 5, pady = 2) # размещение блока

        text = tk.Label(f_top, fg="green", justify=tk.LEFT, padx = 3, pady = 3, font = "Verdana 10") # текст
        ToolTip(text, current_file_name) # имя текущего файла в виде всплывающего окна
        Update_text() # обновление отчёта цифр
        text.pack() # размещение блока

        progress = ttk.Progressbar(f_top, orient = "horizontal", length = 250, mode = 'determinate') # панель прогресса (положение, длина, вид отображения)
        percent_all_failes_number = 100/all_failes_number # перевод в процент от общего числа
        Update_progress() # обновление прогресса
        progress.pack(padx = 4) # размещение блока

        button = tk.Button(f_top, font = "Verdana 11", command = Button_exit, text = "Отмена") # действие кнопки
        button.pack(side = "bottom", pady = 3) # размещение блока

        window.mainloop() # отображение окна

        Stop = True # триггер остановки обработки и сообщения

    Stop = False # триггер остановки сообщения (для работы сообщений при повторном вызове)
    file_number = 0 # отчёт от 0-го файла
    current_file_name = "" # что бы избежать ошибки окна в потоке

    msg_th = Thread(target = Message_count_Thread, args = (all_failes_number, msg, )) # запуск сообщений о количестве файлов в отдельном потоке
    msg_th.start() # запуск потока

    return msg_th # возращаем запущеный поток для определения его завершения

def Change_or_not(n, val, link = None): # запись изменения в последнюю колонку (номер строчки, значение, ссылка)

    cell = ws.cell(row = n, column = ws.max_column) # указание колонки

    if cell.value != val: # если в ячейку не записанно значение

        cell.value = val # запись значения в колонку
        cell.hyperlink = link # гиперссылка на файл

        print(f"Строка {n-1} \"Изменено\": {val}")

def MarkingEmbodimentDocCode(iMarking): # выделение обозначения, исполнения и кода документа, возвращает кортеж (Обозначение, исполнение, пробел перед кодом док., код док.)

    if iMarking != None: # если не надо удалять Обозначение

        docCode =["СБ", "ГЧ", "УЧ", "ЭСБ", "МЧ", "МД", "МС", "Л3"] # список кодов документа

        for docCode in docCode: # перебор кода документа

            whitespaceDocCode = " " + docCode # добавляем пробел перед кодом

            if iMarking.find(whitespaceDocCode)!= -1: # если код документа с пробелом найден
                iMarking = iMarking.replace(whitespaceDocCode, "", 1) # заменить код документа с пробелом на ""
                whitespaceDocCode = " " # пробел перед кодом документа
                break # прервать цикл

            elif iMarking.find(docCode)!= -1: # если код документа найден
                iMarking = iMarking.replace(docCode, "", 1) # заменить код документа на ""
                whitespaceDocCode = "" # без пробела перед кодом документа
                break # прервать цикл

            else: # код документа не подощёл
                continue # использовать следующий

        else: # код документа не найден
            whitespaceDocCode = " " # пробел перед кодом докумена
            docCode = "" # без кода документа

        embodiment = re.findall("-\d{2,3}$", iMarking, re.M) # определение исп. в конце обозначения

        if embodiment: # если найдено исполнение
            embodiment = embodiment[0] # выбираем первый элемент
            iMarking = iMarking.replace(embodiment, "", 1) # заменить исп. на ""
            embodiment = embodiment.replace("-", "",1) # убрать из исп. "-"

        else: # нет исп.
            embodiment = "" # не пишем исп.

    else: # удаляем Обозначение

        iMarking = "" # обозначение
        embodiment = "" # исп.
        whitespaceDocCode = "" # пробел перед кодом документа
        docCode = "" # код докумета

    return iMarking, embodiment, whitespaceDocCode, docCode # возвращаем значения

def Сhange_properties(row, embodimentIndex, iKompasDocument): # изменение св-в документов (список значений из Excel, индекс исполнения файла, интерфейс документа)

    Сhange = False # тригер изменения

    iKompasDocument3D = KompasAPI7.IKompasDocument3D(iKompasDocument) # базовый класс документов-моделей КОМПАС
    iPart7 = iKompasDocument3D.TopPart # интерфейс компонента 3D документа (сам документ)

    if СheckEmbodiment(iPart7, embodimentIndex): # проверка и зменение исп.
        iPart7 = iKompasDocument3D.TopPart # интерфейс компонента 3D документа (сам документ)

    iPropertyKeeper = KompasAPI7.IPropertyKeeper(iPart7) # интерфейс получения/редактирования значения свойств

    for n in range(0, len(row) - 3): # обработка всех столбцов строки кроме последнего

        cell_name = ws.cell(row = 1, column = n + 4) # колонки с заголовками с пропуском первых 3-х
        cell_name = str(cell_name.value).strip() # значение колонок

        cell = str(row[n+2]).strip() # значение из списка

        if cell == "None": # если значение ячейки не записанно
            continue # использовать следующее значение

        elif cell == "delete": # если значение ячейки надо удалить

            if cell_name == "Наименование": # если значение Наименования удалять

                iPart7.Name = "" # прописываем пустоту
                iPart7.Update() # применяем св-ва

                Сhange = True # тригер изменения

                continue # использовать следующее значение

            else: # все остальнвые
                cell = None # пропишем пустое значение

        iGetProperties = iPropertyMng.GetProperties(iKompasDocument) # получить массив св-в

        for iProperty in iGetProperties: # перебор всех св-в из массива

            if cell_name == iProperty.Name.strip(): # сравниваем наименование ячейки с Excel и наименование значения св-ва считаные с документа

                iProperty = iPropertyMng.GetProperty(iKompasDocument, cell_name) # интерфейс свойства
                iPropertyValue = str(iPropertyKeeper.GetPropertyValue(iProperty, 0, True)[1]).strip() # получить значение св-ва (интерфейс св-ва, значение св-ва, единици измерения (СИ))

                if cell != iPropertyValue: # сравниваем значения с Excel и считаные с документа

                    if cell_name == "Обозначение": # если название ячейки "Обозначение"

                        tuple_marking = MarkingEmbodimentDocCode(cell) # выделение обозначения, исполнения и кода документа возвращает кортеж (Обозначение, исполнение, пробел перед кодом док., код док.)

                        СhangeEmbodiment(tuple_marking, iPart7, iKompasDocument3D) # сортировка исполнения (для последовательной записи), присвоение обозначения, исполнения и кода документа

                        Сhange = True # тригер изменения

                        break # прерываем цикл

                    elif cell_name == "Наименование": # если название ячейки "Наименование"

                        iEmbodimentsManager = KompasAPI7.IEmbodimentsManager(iKompasDocument3D) # интерфейс менеджера исп.

                        iEmbodiment = iEmbodimentsManager.Embodiment(0) # получить исп. по индексу или обозначению
                        iPart7 = iEmbodiment.Part # компонент исп.

                        iPart7.Name = cell # прописываем новое наименование
                        iPart7.Update() # применяем св-ва

                        Сhange = True # тригер изменения

                        print(f"{cell_name}: \"{iPropertyValue}\" => \"{cell}\" изменено!")

                        break # прерываем цикл

                    elif cell_name == "Раздел спецификации" and cell != None: # если изменение раздела спецификаци и не удалять значение

                        Сhange_property_SP(iPropertyKeeper, iProperty, cell) # изменение св-в раздела спецификации (интерфейс получения/редактирования значения свойств, интерфейс свойства, значение ячейки)

                        Сhange = True # тригер изменения

                        print(f"{cell_name}: \"{iPropertyValue}\" => \"{cell}\" изменено!!!")

                        break # прерываем цикл

                    elif cell_name == "Плотность": # если название ячейки

                        if cell == None: # если нужно удалить значение
                            cell = 0.0 # прописываем значение "0"

                        else: # если обычное значение
                            cell = float(cell) # значение переводим в вещественое число

                        iPropertyValue = round(float(iPropertyValue), 1) # значение переводим в вещественое число

                        if cell != iPropertyValue: # сравниваем значения с Excel и считаные с документ

                            iSetPropertyValue = iPropertyKeeper.SetPropertyValue(iProperty, cell, True) # установить значение св-ва (интерфейс св-ва, значение св-ва, единици измерения (СИ))
                            iProperty.Update() # применим сво-ва

                            Сhange = True # тригер изменения

                            print(f"{cell_name}: \"{iPropertyValue}\" => \"{cell}\" изменено!")

                        break # прерываем цикл

                    elif cell_name == "Материал": # если название ячейки

                        if cell == None and iPropertyValue == "": # если значение ячейки удалить и прочитанное с-во незаписано

##                            print(f"{cell_name}: \"{cell}\" уже записанно!")

                            break # прерываем цикл

                        else: # если значение нужно менять

                            iSetPropertyValue = iPropertyKeeper.SetPropertyValue(iProperty, cell, True) # установить значение св-ва (интерфейс св-ва, значение св-ва, единици измерения (СИ))
                            iProperty.Update() # применим сво-ва

                            Сhange = True # тригер изменения

                            print(f"{cell_name}: \"{iPropertyValue}\" => \"{cell}\" изменено!")

                            break # прерываем цикл

                    else: # другие св-ва

                        iSetPropertyValue = iPropertyKeeper.SetPropertyValue(iProperty, cell, True) # установить значение св-ва (интерфейс св-ва, значение св-ва, единици измерения (СИ))
                        iProperty.Update() # применим сво-ва

                        Сhange = True # тригер изменения

                        print(f"{cell_name}: \"{iPropertyValue}\" => \"{cell}\" изменено!")

                        break # прерываем цикл

                else: # если свойства совпадают
##                    print(f"{cell_name}: \"{cell}\" уже записанно!")
                    break # прерываем цикл

            else: # если свойство не совпало
                continue # использовать следующее значение

        else: # если св-во не найдено в св-х дет.

            Сhange = Create_property(cell_name, cell, iKompasDocument, iPropertyKeeper) # создать св-во в дет./СБ

    return Сhange # тригер изменения

def СheckEmbodiment(iPart7, embodimentIndex): # проверка и изменение исп.

    iEmbodimentsManager = KompasAPI7.IEmbodimentsManager(iPart7) # интерфейс менеджера исполнений

    iCurrentEmbodimentIndex = iEmbodimentsManager.CurrentEmbodimentIndex # получаем индекс текущего исполнения

    if iCurrentEmbodimentIndex != embodimentIndex: # если текущее исп. отличаеться от заданного
        iSetCurrentEmbodiment = iEmbodimentsManager.SetCurrentEmbodiment(embodimentIndex) # устанавливаем исп.

        return True # возращаем что есть изменения

    else: # текущее исп. не отличаеться от заданного
        return False # возращаем что нет изменения

def СhangeEmbodiment(tuple_marking, iPart7, iKompasDocument3D): # сортировка исполнения (для последовательной записи), присвоение обозначения, исполнения и кода документа

    iEmbodimentsManager = KompasAPI7.IEmbodimentsManager(iKompasDocument3D) # интерфейс менеджера исп.

    iEmbodimentCount = iEmbodimentsManager.EmbodimentCount # узнать количество исп.

    n = iEmbodimentCount - 1 # индекс максимального исп. (отчёт от "0")

    iMarking = tuple_marking[0] # обозначение
    embodiment = tuple_marking[1] # исп.

    if embodiment == "": # если исп. "нулевое"
        embodiment = int("0") # для правильного подсчёта
    else:
        embodiment = int(embodiment) # преобразуем в целое число

    whitespaceDocCode = tuple_marking[2] # пробел перед кодом документа
    docCode = tuple_marking[3] # код докумета

    if iEmbodimentCount > 1: # если исполнений больше 1-го

        iCurrentEmbodiment = iEmbodimentsManager.GetCurrentEmbodimentMarking(2, False) # узнать номер текущего исп. (-1 - всё обозначение; 1 - базовая часть обозначения; 2 - исп. с прочерком; 3 - "1" и "2" вместе; 8 - код документа с пробелом)

        if iCurrentEmbodiment == "" or abs(int(iCurrentEmbodiment)) < embodiment: # если текущее исп. "нулевое" или текущее исп. меньше нового

            embodiment = embodiment + n # максимальный номер исп.

            for n in range(n, -1, -1): # перебор всех исп. в обратном поредке

                iEmbodiment = iEmbodimentsManager.Embodiment(n) # получить исп. по индексу или обозначению
                iPart7 = iEmbodiment.Part # компонент исп.

                RecordMarking(iMarking, embodiment, whitespaceDocCode, docCode, iPart7) # запись обозначения, исп. и кода документа

                embodiment -= 1 # берем исп. на 1 меньше

        else: # текущее исп. не "нулевое"

            if embodiment == "" or abs(int(iCurrentEmbodiment)) >= embodiment: # если новое исп. "нулевое" или текущее исп. больше или равно новому

                for n in range(0, n + 1): # перебор всех исп. в порядке возрастания

                    iEmbodiment = iEmbodimentsManager.Embodiment(n) # получить исп. по индексу или обозначению
                    iPart7 = iEmbodiment.Part # компонент исп.

                    RecordMarking(iMarking, embodiment, whitespaceDocCode, docCode, iPart7) # запись обозначения, исп. и кода документа

                    embodiment += 1 # берем исп. на 1 больше

    else: # только одно исп.
        RecordMarking(iMarking, embodiment, whitespaceDocCode, docCode, iPart7) # запись обозначения, исп. и кода документа

def RecordMarking(iMarking, embodiment, whitespaceDocCode, docCode, iPart7): # запись обозначения, исполнения и кода документа

    if embodiment < 10: # если значение исп. меньше 10
        embodiment = "0" + str(embodiment) # добавляем "0" перед цифрой
    else:
        embodiment = str(embodiment) # преобразуем в строку

    if embodiment == "00": # если в исп. "00"
        embodiment = "" # записать как "нулевое" исп.

    if embodiment != "": # если есть исп.
        dash = "-" # пишем "-"
    else: # нет исп.
        dash = "" # пишем ""

    iCurrentMarking = iPart7.Marking # обозначение текущего исп.
    recording_marking = iMarking + dash + embodiment + whitespaceDocCode + docCode # записываемое обозначение

    if iCurrentMarking != recording_marking: # если обозначения разные

        print(f"Обозначение: \"{iPart7.Marking}\" => \"{recording_marking}\" изменено!")

        full_marking = iMarking + "$|" + dash + "$|" + embodiment + "$|$|$|" + whitespaceDocCode + "$|" + docCode # записываемое обозначение с разделителями

        iPart7.Marking = full_marking # установить обозначение

        iPart7.Update() # применить обозначение

def Сhange_property_SP(iPropertyKeeper, iProperty, cell): # изменение св-в раздела спецификации (интерфейс получения/редактирования значения свойств, интерфейс свойства, значение ячейки)

    dict_xml = {"Сборочные единицы": '<property id="SPCSection" expression="" fromSource="false" format="{$sectionName}">'
                                     '<property id="sectionName" value="Сборочные единицы" type="string" />'
                                     '<property id="sectionNumb" value="15" type="int" />',
                "Детали": '<property id="SPCSection" expression="" fromSource="false" format="{$sectionName}">' # словарь, где значение до ":" это ключ, после значение.
                          '<property id="sectionName" value="Детали" type="string" />'
                          '<property id="sectionNumb" value="20" type="int" />',
                "Стандартные изделия": '<property id="SPCSection" expression="" fromSource="false" format="{$sectionName}">'
                                       '<property id="sectionName" value="Стандартные изделия" type="string" />'
                                       '<property id="sectionNumb" value="25" type="int" />',
                "Прочие изделия": '<property id="SPCSection" expression="" fromSource="false" format="{$sectionName}">'
                                  '<property id="sectionName" value="Прочие изделия" type="string" />'
                                  '<property id="sectionNumb" value="30" type="int" />',
                "Материалы": '<property id="SPCSection" expression="" fromSource="false" format="{$sectionName}">'
                             '<property id="sectionName" value="Материалы" type="string" />'
                             '<property id="sectionNumb" value="35" type="int" />',
                "Комплекты": '<property id="SPCSection" expression="" fromSource="false" format="{$sectionName}">'
                             '<property id="sectionName" value="Комплекты" type="string" />'
                             '<property id="sectionNumb" value="40" type="int" />',
                }

    iSetComplexPropertyValue = iPropertyKeeper.SetComplexPropertyValue(iProperty, dict_xml[cell]) # установить значение св-ва (интерфейс св-ва, значение св-ва, единици измерения (СИ))
    iProperty.Update() # применим сво-ва

def Create_property(cell_name, cell, iKompasDocument, iPropertyKeeper): # создать св-во в дет./СБ

    property_library_path = os.path.join(iSystemPath, property_library) # путь к библиотеке со свойствами

    if os.path.exists(property_library_path): # если есть txt файл использовать его

        iGetProperties = iPropertyMng.GetProperty(property_library_path, cell_name) # интерфейс свойства
        iProperty = iPropertyMng.AddProperty(iKompasDocument, iGetProperties) # создаём св-во

        if iProperty != None: # если есть cв-во в библиотеке

            iPropertyValue = str(iPropertyKeeper.GetPropertyValue(iProperty, 0, True)[1]).strip() # получить значение свойства (интерфейс св-ва, значение св-ва, единици измерения (СИ))
            iSetPropertyValue = iPropertyKeeper.SetPropertyValue(iProperty, cell, True) # установить значение свойства (интерфейс св-ва, значение св-ва, единици измерения (СИ))
            iProperty.Update() # применим сво-ва

            Сhange = True # тригер изменения

            print(f"{cell_name}: \"{iPropertyValue}\" => \"{cell}\" изменено и взято из библиотеки!")

            return Сhange # тригер изменения

        else: # если св-во отсутствует
            Message(f"Отсутствует св-во \"{cell_name}\" в библиотеке \"{property_library_path}\"") # сообщение, поверх всех окон с автоматическим закрытием (текст, время закрытия)
            exit() # выходим из программы

    else:
        Message(f"Не обнаружена библиотека св-в: \"{property_library_path}\"\nРазместите её по указаному пути!") # сообщение, поверх всех окон с автоматическим закрытием (текст, время закрытия)
        exit() # выходим из программы

#-------------------------------------------------------------------------------

list_Excel = [] # список считанных строк из Excel
dict_Settings = {} # словарь считанных настроек из Excel
list_files = [] # список считаных файлов

DoubleExe() # проверка на уже запущеное приложени

KompasAPI() # подключение API компаса

SystemPath() # определяем папку системных файлов

Settings_from_Excel() # настройки считанные с Excel

CheckUpdate() # проверить обновление приложение

Connecting_to_Excel() # подключение EXcel и изменение св-в

Сhecking_open_files() # проверка открытых файлов

Main_assembly() # обработка главной СБ или дет.

Message("Обработка модели закончена!") # сообщение, поверх всех окон с автоматическим закрытием (текст, время закрытия)
